
# ---------------- read from excel file ----------------------------

from openpyxl import Workbook, load_workbook


wb = load_workbook(filename="demoexcel1.xlsx")      # load the workbook which we want to read, create an object of it
sh = wb.active       # open the active sheet of that workbook
# when there is only one sheet in the workbook

print(sh["A5"].value)  # Atul Raman

#----------------------------------------------------------------------------------------


# when there are multiple worksheets in the workbook


wb = load_workbook(filename="demoexcel2.xlsx")
sh = wb.active
# sh = wb["NamePlace"]    # if one of the sheet name is NamePlace

print(sh["A3"].value)  # Supriti

# or

print(wb.active["A4"].value)  # Khushi
print(sh.cell(row=2, column=1).value) #Atul  # not specifing the row, col....but specifying the cell number.


# --------------------------------------------------------------------------------------------

# find max no. of rows and col. in a sheet

wb = load_workbook(filename="demoexcel2.xlsx")
sh = wb.active

rowCount = sh.max_row
colCount = sh.max_column

print(rowCount) # 5
print(colCount) # 2

# accessing all the values form the sheet
for i in range(1, rowCount+1):
    for j in range(1,colCount+1):
        print(sh.cell(row=i,column=j).value, end=" ")  

    print("\n")


# Atul
# Ranchi
# Supriti
# Patna
# Khushi
# Delhi
# Mihir
# Amhadabad


# with end=" " ------------------------

# Atul Ranchi

# Supriti Patna

# Khushi Delhi

# Mihir Amhadabad
