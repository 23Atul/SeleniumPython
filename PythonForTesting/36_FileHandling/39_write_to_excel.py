
# ------------------- write data to excel | openpyxl -----------------------

# https://www.python-excel.org/ ---> modules ( we are using openpyxl )  pip install  openpyxl

# openpyxl is a Python library to read/write Excel 2010 xlsx/xlsm/xltx/xltm files.

# using openpyxl --->   https://openpyxl.readthedocs.io/en/stable/


from openpyxl import Workbook

# simple data writing in excel
wb = Workbook()     # creating workbook object
ws = wb.active      # opening a workbook sheet
ws['A5'] = "Atul Raman"     # workbook cell (r,c)
wb.save("demoexcel1.xlsx")   # saving the workbook

# --------------------------------------------------------------------------


# actual nested data writing in excel

wb = Workbook()
ws = wb.active

testData = [["Name", "city"], ["Atul", "Ranchi"], ["Supriti", "Patna"], ["Khushi", "Delhi"], ["Mihir", "Amhadabad"]]

for data in testData:
    ws.append(data)

wb.save("demoexcel2.xlsx")

#--------------------------------------------------------------------------------------------------------




# filling random data in excel

wb = Workbook()
ws = wb.active

for i in range(1,6):
    for j in range(1,5):
        ws.cell(row=i, column=j).value = i+j * 5    # using cell method

wb.save("demoexcel3.xlsx")





