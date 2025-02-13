
# ---------- Test Data Generation ---------------------------

# Faker is the module in python which helps us to create fake test data for our code or test case

# url - --> https://faker.readthedocs.io/en/latest/index.html

# pip install Faker

# random fake data

from faker import Faker
from openpyxl import Workbook

fakeData = Faker()

print(fakeData.name())  # Margaret Kline
print(fakeData.email())  # christopher79@example.org
print(fakeData.address())  # 8797 Lin Knolls Apt. 996, Deanshire, NJ 67382


#------------------------------------------------------------------------------------

# creating fake data in 1 col. 

# creating fake data and putting it in excel sheet

wb = Workbook()
ws = wb.active

fakeData = Faker()

for i in range(1,11):
    ws.cell(row=i, column=1).value = fakeData.name()
wb.save("testData1.xlsx")

#-----------------------------------------------------------------------

# creating fake data in multiple col

wb = Workbook()
ws = wb.active

fakeData = Faker()

for i in range(1, 11):
    for j in range(1,4):
        ws.cell(row=i, column=1).value = fakeData.name()
        ws.cell(row=i, column=2).value = fakeData.email()
        ws.cell(row=i, column=3).value = fakeData.address()

wb.save("testData2.xlsx")


# --------------------------------------------------------------


# test data with different language (locals methods in same URL)

wb = Workbook()
ws = wb.active

fakeData = Faker(['hi_IN', "en_US"])  # half name in hindi few in english

for i in range(1, 11):
    for j in range(1, 4):
        ws.cell(row=i, column=1).value = fakeData.name()
        ws.cell(row=i, column=2).value = fakeData.email()
        ws.cell(row=i, column=3).value = fakeData.address()

wb.save("testData3.xlsx")
